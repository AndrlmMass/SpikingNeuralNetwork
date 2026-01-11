import argparse
import json
import os
import subprocess
import sys
from pathlib import Path
from typing import List, Tuple, Optional
import math

from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import pandas as pd


DEFAULT_DATASETS = ["MNIST", "KMNIST", "FMNIST"]
DEFAULT_SEEDS = [1, 2, 3, 4, 5]


def run_training_once(dataset: str, seed: int, results_json: Path, extra_args: List[str], sleep_interval: Optional[float] = None) -> Tuple[float, float]:
    """
    Invoke main.py for a single dataset/seed run and return (final_test_acc, final_test_loss).
    Results are read from the provided JSON file that main.py appends to.
    """
    # Ensure results directory exists
    results_json.parent.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable,
        "main.py",
        "--dataset",
        dataset,
        "--runs",
        "1",
        "--results-json",
        str(results_json),
        "--seed",
        str(seed),
        "--no-plot",
    ] + (extra_args or [])
    if sleep_interval is not None:
        cmd += ["--sleep-interval-pct", str(sleep_interval)]

    print(f"\n==> Running dataset={dataset} seed={seed} sleep={sleep_interval}", flush=True)
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    proc = subprocess.run(cmd, cwd=str(Path(__file__).parent), env=env)
    if proc.returncode != 0:
        raise RuntimeError(f"Training failed for dataset={dataset}, seed={seed} (exit={proc.returncode})")

    # Read the last matching record from JSON
    if not results_json.exists():
        raise FileNotFoundError(f"Results JSON not found at {results_json}")
    with open(results_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Results JSON is not a list")

    # Find the most recent matching entry
    last = None
    for entry in reversed(data):
        if str(entry.get("dataset", "")).upper() == dataset.upper() and int(entry.get("seed", -1)) == int(seed):
            if sleep_interval is not None:
                try:
                    si = entry.get("sleep_interval_pct")
                    if si is None:
                        continue
                    if not math.isclose(float(si), float(sleep_interval), rel_tol=1e-9, abs_tol=1e-9):
                        continue
                except Exception:
                    continue
            last = entry
            break
    if last is None:
        raise ValueError(f"No matching results found for dataset={dataset}, seed={seed}, sleep={sleep_interval} in {results_json}")

    final_acc = float(last["final_test_acc"])
    final_loss = float(last["final_test_loss"])
    return final_acc, final_loss


def normalize_label(label: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in str(label)).strip("_")


def _normalize_cols(cols: List[str]) -> List[str]:
    return [normalize_label(c) for c in cols]


def determine_sheet_name(excel_path: Path, desired: Optional[str]) -> str:
    if not excel_path.exists():
        return desired or "Results"
    try:
        xls = pd.ExcelFile(excel_path)
        names = xls.sheet_names
        if desired and desired in names:
            return desired
        # Auto-detect: first sheet with 'model' in headers
        for sn in names:
            try:
                df_head = pd.read_excel(excel_path, sheet_name=sn, nrows=0)
                if "model" in _normalize_cols(list(df_head.columns)):
                    return sn
            except Exception:
                continue
        # Fallback to first sheet
        return desired or names[0]
    except Exception:
        return desired or "Results"


def load_sheet_df(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    if excel_path.exists():
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            # Normalize only logically; keep original columns case for writing back
            return df
        except ValueError:
            # Sheet not present -> create base empty df
            pass
    # Create empty with base headers
    return pd.DataFrame(columns=["model", "run", "Sleep_duration"])


def ensure_df_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    for c in columns:
        if c not in df.columns:
            df[c] = pd.Series([None] * len(df))
    # Keep column order: base keys first, then others (stable)
    base = [c for c in ["model", "run", "Sleep_duration"] if c in df.columns or c in columns]
    others = [c for c in df.columns if c not in base]
    desired_order = base + [c for c in columns if c not in base and c not in others] + others
    df = df.reindex(columns=desired_order)
    return df


def dataset_seed_column_name(dataset: str, seed: int) -> str:
    return f"{normalize_label(dataset)}_s{int(seed)}"


def ensure_df_dataset_seed_columns(df: pd.DataFrame, datasets: List[str], seeds: List[int]) -> pd.DataFrame:
    needed = [dataset_seed_column_name(ds, s) for ds in datasets for s in seeds]
    return ensure_df_columns(df, needed)


def find_or_create_row_idx(df: pd.DataFrame, model: str, run_index: int, sleep_duration: float) -> int:
    # Try exact match
    mask = (
        (df.get("model").astype(str).str.strip().str.lower() == str(model).strip().lower()) &
        (pd.to_numeric(df.get("run"), errors="coerce") == int(run_index)) &
        (pd.to_numeric(df.get("Sleep_duration"), errors="coerce") == float(sleep_duration))
    )
    idxs = df.index[mask].tolist()
    if idxs:
        return idxs[0]
    # Append new row
    new_row = {"model": model, "run": int(run_index), "Sleep_duration": float(sleep_duration)}
    df.loc[len(df)] = new_row
    return len(df) - 1


def save_df_to_excel(df: pd.DataFrame, excel_path: Path, sheet_name: str) -> None:
    # Create directories if needed
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    # Preserve other sheets if workbook exists
    if excel_path.exists():
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            writer.book  # ensure book loaded
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def get_header_map(ws) -> dict:
    header = {}
    for col_idx, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        val = cell[0].value
        if val is None:
            continue
        header[normalize_label(val)] = col_idx
    return header


def ensure_dataset_seed_columns(ws, datasets: List[str], seeds: List[int]) -> dict:
    """
    Ensure header row has columns for each dataset and seed named <dataset>_s<seed>,
    e.g., 'mnist_s1', 'mnist_s2', ... Returns mapping {(dataset_upper, seed) -> column_index}.
    """
    header = get_header_map(ws)
    col_map = {}
    next_col = ws.max_column + 1 if ws.max_column else 1
    for ds in datasets:
        ds_key = normalize_label(ds)
        for seed in seeds:
            key = f"{ds_key}_s{int(seed)}"
            if key in header:
                col_map[(ds.upper(), int(seed))] = header[key]
            else:
                ws.cell(row=1, column=next_col, value=key)
                col_map[(ds.upper(), int(seed))] = next_col
                next_col += 1
    return col_map


def find_or_create_row_by_model_and_run(ws, model_name: str, run_index: int) -> int:
    header = get_header_map(ws)
    model_col = header.get("model")
    run_col = header.get("run")
    if model_col is None:
        model_col = 1
        ws.cell(row=1, column=model_col, value="model")
    if run_col is None:
        run_col = 2
        ws.cell(row=1, column=run_col, value="run")

    model_target = str(model_name).strip().lower()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        model_cell = row[model_col - 1]
        run_cell = row[run_col - 1]
        if model_cell.value is None or run_cell.value is None:
            continue
        if str(model_cell.value).strip().lower() == model_target and int(run_cell.value) == int(run_index):
            return model_cell.row

    # Not found: append new row
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=model_col, value=model_name)
    ws.cell(row=new_row, column=run_col, value=int(run_index))
    return new_row
def find_or_create_row_by_model_run_sleep(ws, model_name: str, run_index: int, sleep_duration: float) -> int:
    header = get_header_map(ws)
    model_col = header.get("model") or 1
    run_col = header.get("run") or 2
    sleep_col = header.get("sleep_duration")
    if header.get("model") is None:
        ws.cell(row=1, column=model_col, value="model")
    if header.get("run") is None:
        ws.cell(row=1, column=run_col, value="run")
    if sleep_col is None:
        sleep_col = max(model_col, run_col, ws.max_column or 2) + 1
        ws.cell(row=1, column=sleep_col, value="Sleep_duration")

    model_target = str(model_name).strip().lower()

    def to_float(x):
        try:
            return float(x)
        except Exception:
            return None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        model_cell = row[model_col - 1]
        run_cell = row[run_col - 1]
        sleep_cell = row[sleep_col - 1]
        if model_cell.value is None or run_cell.value is None or sleep_cell.value is None:
            continue
        if str(model_cell.value).strip().lower() != model_target:
            continue
        if int(run_cell.value) != int(run_index):
            continue
        sv = to_float(sleep_cell.value)
        if sv is None:
            continue
        if abs(sv - float(sleep_duration)) < 1e-12:
            return model_cell.row

    # Not found: append new row
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=model_col, value=model_name)
    ws.cell(row=new_row, column=run_col, value=int(run_index))
    ws.cell(row=new_row, column=sleep_col, value=float(sleep_duration))
    return new_row


def write_result(ws, row_idx: int, col_idx: int, value: float) -> None:
    ws.cell(row=row_idx, column=col_idx, value=float(value))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=str, default=str(Path(__file__).parent / "Results_.xlsx"),
                        help="Path to Excel workbook to update")
    parser.add_argument("--datasets", type=str, nargs="*", default=DEFAULT_DATASETS,
                        help="Datasets to iterate (e.g., MNIST KMNIST FMNIST)")
    parser.add_argument("--seeds", type=int, nargs="*", default=DEFAULT_SEEDS,
                        help="Seed values to use per dataset")
    parser.add_argument("--model-name", type=str, default="snntorch",
                        help="Model row name in column A")
    parser.add_argument("--results-json", type=str, default=str(Path(__file__).parent / "results" / "orchestrate_runs.json"),
                        help="JSON file where runs append results (separate from default)")
    parser.add_argument("--sleep-intervals", type=float, nargs="*", default=[0.1],
                        help="One or more sleep interval fractions (e.g., 0.05 0.1 0.2)")
    parser.add_argument("--sheet", type=str, default=None,
                        help="Target worksheet name to update (default: autodetect by headers)")
    parser.add_argument("--log-updates", action="store_true",
                        help="Print a concise summary of each DataFrame update after saving")
    parser.add_argument("--extra", type=str, nargs=argparse.REMAINDER, default=[],
                        help="Extra args to pass to main.py after '--'")
    args = parser.parse_args()

    # Handle optional '--' for extra args
    extra_args = []
    if args.extra:
        extra_args = [a for a in args.extra if a != "--"]

    results_json = Path(args.results_json)

    excel_path = Path(args.excel).resolve()
    target_sheet = determine_sheet_name(excel_path, args.sheet)
    print(f"Updating workbook: {excel_path} | sheet: {target_sheet}", flush=True)
    df = load_sheet_df(excel_path, target_sheet)
    df = ensure_df_columns(df, ["model", "run", "Sleep_duration"])
    df = ensure_df_dataset_seed_columns(df, args.datasets, args.seeds)

    # Always write to run=1 row as requested; vary seed via columns
    run_index = 1
    total_runs = len(args.sleep_intervals) * len(args.datasets) * len(args.seeds)
    with tqdm(total=total_runs, desc="Orchestrating", leave=True) as pbar:
        for sleep in args.sleep_intervals:
            for ds in args.datasets:
                for seed in args.seeds:
                    acc, loss = run_training_once(ds, seed, results_json, extra_args, sleep_interval=sleep)
                    row_idx = find_or_create_row_idx(df, args.model_name, run_index, float(sleep))
                    col_name = dataset_seed_column_name(ds, int(seed))
                    if col_name not in df.columns:
                        df = ensure_df_columns(df, [col_name])
                    df.at[row_idx, col_name] = float(acc)
                    try:
                        save_df_to_excel(df, excel_path, target_sheet)
                    except PermissionError as e:
                        print(f"PermissionError saving workbook. Is it open in Excel? {e}", flush=True)
                        raise
                    if args.log_updates:
                        try:
                            snap = df.loc[row_idx, ["model", "run", "Sleep_duration", col_name]].to_dict()
                        except Exception:
                            snap = {"model": args.model_name, "run": 1, "Sleep_duration": float(sleep), col_name: float(acc)}
                        print(f"DF update -> row={row_idx} {snap}", flush=True)
                    pbar.set_postfix({"dataset": ds, "seed": int(seed), "sleep": float(sleep), "acc": f"{acc:.4f}"})
                    pbar.update(1)

    try:
        save_df_to_excel(df, excel_path, target_sheet)
    except PermissionError as e:
        print(f"PermissionError saving workbook. Is it open in Excel? {e}", flush=True)
        raise
    print(f"Updated workbook: {excel_path}", flush=True)


if __name__ == "__main__":
    main()


